from openpyxl import load_workbook

def main():
    # 1️⃣ Load existing Excel file
    wb = load_workbook("sales_data.xlsx")

    # 2️⃣ Select sheet "2025"
    ws = wb["2025"]

    # 3️⃣ Add header for Total column
    ws["D1"] = "Total"

    # 4️⃣ Loop through rows and calculate total
    for row in range(2, ws.max_row + 1):
        quantity = ws.cell(row=row, column=2).value  # Column B
        price = ws.cell(row=row, column=3).value     # Column C
        total = quantity * price
        ws.cell(row=row, column=4, value=total)      # Column D

    # 5️⃣ Save to new Excel file
    wb.save("sales_summary.xlsx")

    print("sales_summary.xlsx created successfully!")

if __name__ == "__main__":
    main()
